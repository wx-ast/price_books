import os
import csv
import re
import io
import importlib

import xlrd
from openpyxl import load_workbook
import xlwt
from server_timing.middleware import timed_wrapper, TimedService, timed

from django.shortcuts import render, redirect
from django.core.files.storage import FileSystemStorage
from django.views.generic import View
from django.views.generic.base import ContextMixin, TemplateResponseMixin
from django.http import HttpResponse
from django.conf import settings
from django.urls import reverse

from .forms import PriceForm, OrderForm
from .models import Product, Supplier, Order, OrderItem
from .utils import process_order_item


class FileUploadBaseView(TemplateResponseMixin, ContextMixin, View):
    template_name = 'books/load_file.html'
    form_cls = PriceForm
    started = False
    loader = None

    def __init__(self, *args, **kwargs):
        super(FileUploadBaseView, self).__init__(*args, **kwargs)
        self.form = self.form_cls()

    def post(self, request):
        post_service = TimedService('post')
        post_service.start()

        if not request.FILES:
            data = {
                'error': True,
                'error_text': 'file not posted'
            }
            return self.render_to_response(self.get_context_data(**data))

        form = self.form_cls(request.POST, request.FILES)
        if not form.is_valid():
            data = {
                'error': True,
                'error_text': 'form not valid'
            }
            return self.render_to_response(self.get_context_data(**data))

        if not self.process_post_data(form.cleaned_data):
            data = {
                'error': True,
                'error_text': 'что-то пошло не так'
            }
            return self.render_to_response(self.get_context_data(**data))

        postfile = request.FILES['file']
        filename, file_extension = os.path.splitext(postfile.name)
        if file_extension not in ('.xls', '.xlsx'):
            data = {
                'form': form,
                'error': True,
                'error_text': 'Поддерживаются только *.xls, *.xlsx'
            }
            return self.render_to_response(self.get_context_data(**data))

        fs_storage = FileSystemStorage()
        filename = fs_storage.save(postfile.name, postfile)
        uploaded_file_path = fs_storage.path(filename)

        if file_extension == '.xls':
            workbook = xlrd.open_workbook(
                uploaded_file_path,
                formatting_info=False
            )
            sheet = workbook.sheet_by_index(0)
            for rownum in range(sheet.nrows):
                row = sheet.row_values(rownum)
                if not self.process_line(row):
                    break
        elif file_extension == '.xlsx':
            workbook = load_workbook(
                filename=uploaded_file_path,
                read_only=True
            )
            worksheet = workbook.active
            for row in worksheet.rows:
                row = [cell.value for cell in row]
                if not self.process_line(row):
                    break

        # uploaded_file = open(uploaded_file_path, encoding='cp1251')
        # csvfile = csv.reader(uploaded_file, delimiter=';', quotechar='"')
        # for row in csvfile:
        #     if not self.process_line(row):
        #         break

        if hasattr(self.loader, 'bindings') and \
                self.loader.bindings is not None:
            print(self.loader.bindings)

        # uploaded_file.close()
        os.remove(uploaded_file_path)

        data = {
            'message': 'Загрузка завершена'
        }
        post_service.end()
        with timed('render'):
            return self.render_to_response(self.get_context_data(**data))

    def get(self, *args, **kwargs):
        data = {
            'form': self.form
        }
        return self.render_to_response(self.get_context_data(**data))

    def process_post_data(self, data):
        print(data)
        return True

    def process_line(self, row):
        print(row)
        return True


class PriceLoadView(FileUploadBaseView):

    def process_post_data(self, data):
        supplier = Supplier.objects.get(
            pk=data.get('supplier'))
        Product.objects.filter(supplier=supplier).delete()

        try:
            loader_module = importlib.import_module(
                f'.loaders.{supplier.loader_type}_loader',
                package=__package__
            )
            self.loader = loader_module.Loader(supplier)
        except ModuleNotFoundError as e:
            print(e)
            return False
        return True

    def process_line(self, row):
        return self.loader.process_line(row)


class OrderLoadView(FileUploadBaseView):
    form_cls = OrderForm
    i = 0

    def process_post_data(self, data):
        order = Order(name=data['file'].name)
        order.save()
        order_loader = data.get('loader', 'df')

        try:
            loader_module = importlib.import_module(
                f'.loaders.{order_loader}_order_loader',
                package=__package__
            )
            self.loader = loader_module.OrderLoader(order)
        except ModuleNotFoundError as e:
            print(e)
            return False
        return True

    def process_line(self, row):
        return self.loader.process_line(row)


@timed_wrapper('index')
def index(request):
    home_service = TimedService('select')
    home_service.start()

    orders = Order.objects.all()
    items = []
    for order in orders:
        data = {
            'order': order,
            'count': {},
            'count_all': OrderItem.objects.filter(order=order).count()
        }
        for i in range(4):
            data['count'][i] = OrderItem.objects.\
                filter(order=order, status=i).count()

        items.append(data)
    home_service.end()
    context = {
        'items': items
    }

    with timed('render'):
        return render(request, 'books/index.html', context)


def order_table(request, order_id, status):
    items = OrderItem.objects.filter(order=order_id, status=status)

    context = {
        'items': items
    }
    return render(request, 'books/order_table.html', context)


def get_csv_file(request, order_id, status):
    items = OrderItem.objects.filter(order=order_id, status=status)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = \
        f'attachment; filename="order_{status}.csv"'

    writer = csv.writer(response, delimiter=';', quotechar='"')
    for item in items:
        row = [item.count, item.name, item.author, item.binding,
               item.publisher, item.quantity]
        if item.product:
            row.append(item.product.article)
            row.append(item.product.author)
            row.append(item.product.publisher)
            row.append(item.product.binding)
            row.append(item.product.price)
            row.append(item.product.supplier.name)
        writer.writerow(row)

    return response


def get_xls_file(request, order_id, status):
    items = OrderItem.objects.filter(order=order_id, status=status)

    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet(f'order_{status}')

    i = j = 0
    for item in items:
        row = [item.count, item.name, item.author, item.binding,
               item.publisher, item.quantity]
        if item.product:
            row.append(item.product.article)
            row.append(item.product.author)
            row.append(item.product.publisher)
            row.append(item.product.binding)
            row.append(item.product.price)
            row.append(item.product.supplier.name)

        for cell in row:
            worksheet.write(i, j, cell)
            j += 1
        i += 1
        j = 0

    file_path = os.path.join(settings.MEDIA_ROOT, f'order_{status}.xls')
    workbook.save(file_path)
    xlsfile = open(file_path, 'rb')
    file_data = xlsfile.read()
    xlsfile.close()

    response = HttpResponse(file_data, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = \
        f'attachment; filename="order_{status}.xls"'
    return response


def order_update(request, order_id):
    order = Order.objects.get(pk=order_id)
    items = OrderItem.objects.filter(order=order)
    for item in items:
        data = {
            'name': item.name,
            'author': item.author,
            'binding': item.binding,
            'publisher': item.publisher,
            'quantity': item.quantity,
        }
        ret, ret_data = process_order_item(data, order)
        if ret:
            item.product = ret_data['product']
            item.status = ret_data['status']
            item.count = ret_data['count']
        else:
            item.product = None
            item.status = 0
            item.count = 0
        item.save()

    return render(request, 'books/index.html')


def order_delete(request, order_id):
    order = Order.objects.get(pk=order_id).delete()

    return redirect(reverse('index'))
